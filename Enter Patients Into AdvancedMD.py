##############################################
#All rights reserved Copy Rights 2022 Ashton Fox.
#This code cannot be used to sell or make money with in any way, shape, or form.
##############################################

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from time import sleep
from Patient import Patient

#The reason there is a Debug Mode is I purposely added dealys between each key being typed when I send input to a form input so I would not look like a bot, but this just takes forever to debug
IsInDebugMode = False

AlertWaitTime = 0.2

delay = 30

driver = webdriver.Chrome(ChromeDriverManager().install())

def CheckForPopups():
   
   #Checking multiple times because sometimes there are multiple popups
   try:
         sleep(AlertWaitTime)
         driver.switch_to_alert().accept()
     except:
         pass
     try:
         sleep(AlertWaitTime)
         driver.switch_to_alert().accept()
     except:
         pass
     try:
         sleep(AlertWaitTime)
         driver.switch_to_alert().accept()
     except:
         pass
     try:
         sleep(AlertWaitTime)
         driver.switch_to_alert().accept()
     except:
         pass
     try:
         sleep(AlertWaitTime)
         driver.switch_to_alert().accept()
     except:
         pass
     try:
         sleep(AlertWaitTime)
         driver.switch_to_alert().accept()
     except:
         pass
def SavePatient():
   #This might look a little weired but the software would not allow us to just click the save button to save the patients, it made you use the shortcut ALT + A to save. so i had to get kind of clever and send the keys Alt and A to the body element of the page for it to work
   elem = driver.find_element_by_xpath("/html/body")
   elem.send_keys(Keys.ALT, "a") 
def ClearForm():
   #Again this might look a little weired but as i said before the software would not allow us to just click the clear button to clear the form, it made you use the shortcut ALT + L to clear the form. so i had to get kind of clever and send the keys Alt and L  to the body element of the page for it to work
   elem = driver.find_element_by_xpath("/html/body")
   elem.send_keys(Keys.ALT, "l") 
#This function is here so if it takes a little bit longer to load, my code does not fail
#i do this by just trying to find the element i want and then seeing if there is an error or not. if so i try it again... until eventualy it works.
def WaitForElement(Element, FindBy):
        IsLoaded = False
        while IsLoaded == False:

            if FindBy.upper() == "ID":
                 try:
                        driver.find_element(By.ID, str(Element))
                        IsLoaded = True
                        break
                 except:
                    continue
            elif FindBy.upper() == "XPATH":
                try:
                        driver.find_element(By.XPATH, str(Element))
                        IsLoaded = True
                        break
                except:
                    continue
            elif FindBy.upper() == "CSS":
                try:
                        driver.find_element(By.CSS_SELECTOR,str(Element))
                        IsLoaded = True
                        break
                except:
                    continue
            elif FindBy.upper() == "NAME":
                try:
                    driver.find_element(By.Name,str(Element))
                    IsLoaded = True
                    break
                except:
                    continue
                     
#This is here because there was a diffrent way you had to login to the software when you where entering the patients "Personal Data" Like Street Address, Email Address, DOB; insted of just entering the Patients Insurance Information like their Insurance Policy Number. 
def LoginForPersonalData(username, password, officeKey):
    driver.get("https://login.advancedmd.com/")
    driver.switch_to.frame(0)
    driver.find_element(By.NAME, "loginname").send_keys(username)
    driver.find_element(By.NAME, "password").send_keys(password)
    driver.find_element(By.NAME, "officeKey").send_keys(officeKey)
    driver.find_element(By.CSS_SELECTOR, ".btn").click()
    sleep(12)
    window_after = driver.window_handles[1]
    driver.minimize_window()
    driver.switch_to.window(window_after)
    driver.maximize_window()
    sleep(AlertWaitTime)
    WebDriverWait(driver, delay).until(EC.element_to_be_clickable((By.ID, 'mnuPatientInfo'))).click()
    sleep(AlertWaitTime)
    driver.switch_to.frame(0)
    sleep(6)
    driver.switch_to.frame(0)
      
 # As i said in the previous comment this is here because there was a diffrent way you had to login to the medical billing software depending on if you where entiring patient personal data, or insurance information
def LoginForInsuranceData(username, password, officeKey):
    driver.get("https://login.advancedmd.com/")

    driver.switch_to.frame(0)
    driver.find_element(By.NAME, "loginname").send_keys(username)
    driver.find_element(By.NAME, "password").send_keys(password)
    driver.find_element(By.NAME, "officeKey").send_keys(officeKey)
    WebDriverWait(driver, delay).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn"))).click()
    sleep(30)
    window_after = driver.window_handles[1]
    driver.minimize_window()
    driver.switch_to.window(window_after)
    driver.maximize_window()

    WebDriverWait(driver, delay).until(EC.element_to_be_clickable((By.ID, 'mnuPatientInfo'))).click()
    sleep(3)
    driver.switch_to.frame(0)
    WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.XPATH, '//*[@id="cdk-drop-list-0"]/div[4]'))).click()
    driver.switch_to.frame(0)
#this function is what actually enters the patients personal information into the form
def EnterNewPatientPersonal(patient):
   CheckForPopups()
    if len(patient.PatientZipCode.strip()) == 4:
        patient.PatientZipCode = ("0" + patient.PatientZipCode)
    #Clearing The Form And Makeing Sure It Is In The Patint Info Tab
    WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.TAG_NAME, "body"))).send_keys(Keys.ALT, "l")

    #Creating Varibles For Elements On Patient Information Page
    PatientAdditionalMrn_Input = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.ID, 'txtPtAddtMRN')))
    PatientFullName_Input = driver.find_element(By.CSS_SELECTOR, "#ellPtPatient > input")
    PatientStreetAddress_Input = driver.find_element(By.ID, "txtPtAddress2")
    PatientZipCode_Input = driver.find_element(By.ID, "ellPtZipCode_textbox")
    Gender_Dropdown = driver.find_element(By.ID, "selPtGender").find_element(By.XPATH, "//option[. = '"+patient.PatientGender.upper().strip()+"']")
    PatientDOB_Input = driver.find_element(By.ID, "txtPtDOB")
    PatientAccountNumberInput = driver.find_element(By.ID, "txtPtAddtMRN")
    ResponsablePartyInput = driver.find_element(By.CSS_SELECTOR, "#ellPtRespParty > input[type=text]")

    #Clearing Form
    PatientAdditionalMrn_Input.clear()
    PatientFullName_Input.clear()
    PatientStreetAddress_Input.clear()
    PatientZipCode_Input.clear()
    PatientDOB_Input.clear()
    PatientAccountNumberInput.clear()
    ResponsablePartyInput.clear()

    #Filling Out Patient Info Form
    PatientFullName_Input.send_keys(patient.PatientFullName)
    ResponsablePartyInput.send_keys(patient.PatientFullName)
    PatientStreetAddress_Input.send_keys(patient.PatientStreetAddress)
    PatientZipCode_Input.send_keys(patient.PatientZipCode, Keys.TAB)
    
    PatientAccountNumberInput.send_keys(patient.PatientAccountNumber)
    Gender_Dropdown.click()
    PatientDOB_Input.send_keys(patient.PatientDOB)
    sleep(AleartWaitTime)
    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ALT, "a")
      
    CheckForPopups()

#This function enters the patients Insurance Data
def EnterNewPatientInsurance(patient):
    print(patient.PatientFullName)

    driver.find_element(By.ID, "txtInsEndDate").clear()
    driver.switch_to.parent_frame()

    #Searching For Patient
    driver.find_element(By.ID, "mat-input-0").clear()
    driver.find_element(By.ID, "mat-input-0").click()
    driver.find_element(By.ID, "mat-input-0").send_keys(patient.PatientFullName)
    WebDriverWait(driver, delay).until(EC.element_to_be_clickable((By.TAG_NAME, 'mat-option'))).click()
    
    #Switching To The Form Frame
    driver.switch_to.frame(0)
    sleep(0.3)

    #Defining Varibles For Elements
    SubscriberID_Input = driver.find_element(By.ID, "txtSubScriberIDNumber")
    Carrier_Input = driver.find_element(By.XPATH, '//*[@id="ellCarrier"]/input')
    EffectiveDate_Input = driver.find_element(By.XPATH, '//*[@id="txtInsBeginDate"]')

    #Clearing The Form
    SubscriberID_Input.clear()
    Carrier_Input.clear()
    EffectiveDate_Input.clear()

    CheckForPopups()
   
    Carrier_Input.send_keys(patient.Carrier, Keys.TAB)
    #Waiting For The Carrier Input To Load
    sleep(1)
    CheckForPopups()
         
    #Filling Out The Form
    SubscriberID_Input.send_keys(patient.PatientInsID)
    EffectiveDate_Input.send_keys("01012020")
    sleep(0.3)
    driver.find_element(By.ID, "btnSave").click()
    ActionChains(driver).key_down(Keys.ALT).send_keys('a').key_up(Keys.ALT).perform()
    sleep(0.3)
    
   CheckForPopups()
         
 #sometimes it is neccasary to actually delete patients, thats what this function dose
def DeletePatient(patient):
    driver.switch_to.parent_frame()
    #Searching For Patient 
    sleep(0.1)
    WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.ID, "mat-input-0"))).clear()
    WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.ID, "mat-input-0"))).click()
    WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.ID, "mat-input-0"))).send_keys(patient)
    sleep(1)
    try:
        driver.find_element(By.TAG_NAME, "mat-option").click()
        sleep(0.3)
        driver.switch_to.frame(0)
        sleep(0.3)
        driver.find_element(By.ID, "btnPtDelete").click()
        sleep(0.2)
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ALT, "t")
        sleep(0.2)
      
       CheckForPopups()
        
    except:
        element = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]')
        driver.execute_script("""
        var element = arguments[0];
        element.parentNode.removeChild(element);
        """, element)
        sleep(0.1)
        driver.find_element(By.XPATH, '/html/body/amds-root/div/amds-layout/mat-drawer-container/mat-drawer-content/amds-header-toolbar/mat-toolbar/div[1]/amds-patient-lookup-control/div/mat-form-field/div/div[1]/div[3]/button[1]').click()
        driver.switch_to.frame(0)
        return
#This is the main function that gets the patients information from an Excel Spreadsheet and then enters that info into AdvancedMD
def EnterPatientsFromList(PatientListFileName, username, password, officeKey, IsEnteringForPatientPersonalData=True, SheetName='Sheet1', StartAt=2, IsClearingPatients=False):
    if IsClearingPatients:
        LoginForPersonalData(username, password, officeKey)
        book = load_workbook(filename = str(PatientListFileName).strip())
        ws = book[SheetName]
        rows = ws.iter_rows(min_row=2, min_col=1, max_col=2)
        for FullName, Blank in rows:
            #NewPatient = Patient(" ", " ", " ", " ", " ", " ", "(XXX) XXX-XXXX", " ", " ", "COMMERCIAL", " ", " ", " ")
            print("'" + FullName.value + "'")
            #NewPatient = Patient(FirstName.value, LastName.value, Gender.value, DOB.value, Email.value, PrimaryPhone.value, "(XXX) XXX-XXXX", StreetAddrs.value, ZipCode.value, "COMMERCIAL", InsuranceID.value, AccountNumber.value, Carrier.value)
            DeletePatient(FullName.value)
        driver.close()
    elif IsEnteringForPatientPersonalData:
        LoginForPersonalData(username, password, officeKey)
        book = load_workbook(filename = str(PatientListFileName).strip())
        ws = book[SheetName]
        if IsInDebugMode:
            rows = ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=15)
        else:
            rows = ws.iter_rows(min_row=StartAt, min_col=1, max_col=15)
        for FullName, FirstName, LastName, StreetAddrs, City, State, ZipCode, PrimaryPhone, Email, Gender, DOB, AccountNumber, Provider, InsuranceID, Carrier in rows:
            NewPatient = Patient(FirstName.value, LastName.value, Gender.value, DOB.value, Email.value, PrimaryPhone.value, "(XXX) XXX-XXXX", StreetAddrs.value, ZipCode.value, "COMMERCIAL", InsuranceID.value, AccountNumber.value, Carrier.value)
            EnterNewPatientPersonal(NewPatient)
        driver.close()
    else:
        LoginForInsuranceData(username, password, officeKey)
        book = load_workbook(filename = str(PatientListFileName).strip())
        ws = book[SheetName]
        if IsInDebugMode:
            rows = ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=15)
        else:
            rows = ws.iter_rows(min_row=StartAt, min_col=1, max_col=15)
        for FullName, FirstName, LastName, StreetAddrs, City, State, ZipCode, PrimaryPhone, Email, Gender, DOB, AccountNumber, Provider, InsuranceID, Carrier in rows:
             NewPatient = Patient(FirstName.value, LastName.value, Gender.value, DOB.value, Email.value, PrimaryPhone.value, "(XXX) XXX-XXXX", StreetAddrs.value, ZipCode.value, "COMMERCIAL", InsuranceID.value, AccountNumber.value, Carrier.value)
             EnterNewPatientInsurance(NewPatient)
        driver.close()
